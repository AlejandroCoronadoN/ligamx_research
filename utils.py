import pandas as pd
import logging 
import numpy as np
import datetime 
import pdb 
from sklearn import linear_model
from openpyxl import load_workbook

def create_unique_enfrentamiento(df, team_list):
  """This function is used to create a unique enfrentamiento tag for 
  all the possible combinations of equipo_visitante and equipo_local
  such that for e' in equipo_visitante and e'' in equipo_local
   efrentamiento(e',e'')== enfrentamiento(e'',e')

  Args:
      df (DataFrame): df_matchhes, BBVA dataset
      team_list ([string]): List with the name of all the teams
  """
  enfrentamiento_list =[]
  for e in team_list:
    for e_2 in team_list:
      if e == e_2:
        continue
      else:
        enfrentamiento_list.append(e + " VS " +e_2 )
  
  df['enfrentamiento'] = df['equipo_local'] + " VS " + df['equipo_visitante']
  for enfrentamiento in enfrentamiento_list:
    team1 = enfrentamiento.split(" VS ")[0]
    team2 = enfrentamiento.split(" VS ")[1]
    df.loc[df['enfrentamiento'].apply(lambda x: True if (team1 == x.split(" VS ")[0] or team1 == x.split(" VS ")[1]) and (team2 == x.split(" VS ")[0] or team2 == x.split(" VS ")[1]) else False), 'enfrentamiento'] = enfrentamiento

  return df

def round_hours(df, col):
  """This function uses datetime.time variables and round them to
  complete or half hours. In the context of the dataset there are only a couple
  of matches that start at fraction hours. We would like to condense the information
  in bigger blocks.

  Args:
      df ([DataFrame]): []
      col ([string]): Name of the column to perform operation
  #TODO Create test for parsing hours correctly, what if we don't get datetime.time variables?
  """
  map_blocks = {}
  time_blocks = df[col].unique()
  col_round = col + '_round'
  df[col_round] = df[col].copy()
  for block in time_blocks:
    #pdb.set_trace()
    if block.minute !=0 and block.minute != 30:
      if 30-block.minute  > 15: #round to 0:
        newblock = datetime.time(block.hour, 0)
      else: #round to 30
        newblock = datetime.time(block.hour, 30)
      df.loc[df[col]==block, col_round] = newblock
  
  return df

def read_distance_file(path, sheet, r1, r2):
  """Read the distance matrix in Asistencia BBVA. It selects only the 
  cells inside the range [r1,r2]. Then we prepare the headers for this 
  matrix a insert the index. The returning DataFram is a symetric matrix 
  with each row and column indicating the name of the team.

  Args:
      path (string): Name of the  path
      sheet (string): Excel Sheet name
      r1 (string): Range 1 Ex: "B2"
      r2 (string): Range 2 Ex: "AB100"

  Returns:
      [type]: 
  """
  wb = load_workbook(filename=path, read_only=True, data_only=True)
  ws = wb[sheet]
  # Read the cell values into a list of lists
  data_rows = []
  for row in ws[r1:r2]:
      data_cols = []
      for cell in row:
          data_cols.append(cell.value)
      data_rows.append(data_cols)

  df = pd.DataFrame(data_rows)
  df.columns = df.loc[0]
  df = df.drop([0])
  df = df.set_index(df[df.columns[0]])
  del df[df.columns[0]]
  df.index.name=None
  return df


#Replace missing information for numeric variables
def missingnumeric_regression_autocomplete(df, targetvar):
  """ Makes a regression for each missing value for the numeric columns
  using fixed variables

  Args:
      df ([type]): [description]
      targetvar (string): traget variable.
  """
  independent_variables = ['aforo', 'estadio_asistencia', 'goles_anotados']
  categorical_variables = ['estadio', 'enfrentamiento', 'dia_semana', 'hora_bloque']
  df['hora_bloque'] = df.hora_bloque
  
  X = df.copy()
  missing_values =  X.loc[X[targetvar].isna(),]

  regression_columns = [x for x in X.columns if (x in independent_variables or x in categorical_variables or  x == targetvar)]
  X = X[regression_columns]

  for catvar in categorical_variables:
    if np.sum(missing_values[catvar].isna())==0:
      logging.debug('missingnumeric_regression_autocomplete** pd.get_dummies: '+ catvar)
      X = pd.concat([X, pd.get_dummies(X[catvar])], axis=1)

    del X[catvar] 
  missing_values =  X.loc[X[targetvar].isna(),]

    
  for var in independent_variables:
    if np.sum(missing_values[var].isna())==0:
      missing_values[var] = pd.to_numeric(missing_values[var])
      X[var] = pd.to_numeric(X[var])
    elif var == targetvar:
      X=X[~ X[targetvar].isna()]
    else: 
      del X[var]
      del missing_values[var]
      independent_variables.remove(var)
      
  X = X.dropna()      
  y = X[targetvar]
  del X[targetvar]
  reg = linear_model.LinearRegression().fit(X, y) 
  results = reg.predict(missing_values[X.columns])
  missing_values[targetvar] =  results
  #Replace predicted values to the original dataset
  df.loc[missing_values.index, targetvar] = results
  return df
  
def fill_empty_distances(df, df_distance):
  df_empty = df[df['distancia_equipos'].sina()]
  for i in df_empty.index:
    equipo_visitante= df.loc[i, 'equipo_visitante']
    equipo_local = df.loc[i, 'equipo_local']
    df_distance.loc[equipo_visitante, equipo_local]
    